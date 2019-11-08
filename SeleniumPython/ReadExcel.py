import openpyxl

path="D:\SeleniumPython\dummyfiles\sample.xlsx"

workbook=openpyxl.load_workbook(path)
#sheet=workbook.active
#sheet=workbook.worksheets[0]
sheet=workbook["Sheet1"]

rows=sheet.max_row
cols=sheet.max_column

print("No of Rows: ",rows)
print("No of Columns :",cols)

print(sheet.cell(row=5,column=6).value)


for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end=" | ")
        #print(sheet.cell(row=r, column=c).value)
    print()