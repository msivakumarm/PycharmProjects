import openpyxl
import random

path="D:\SeleniumPython\dummyfiles\Test.xlsx"
workbook=openpyxl.load_workbook(path)
#sheet=workbook.active
#sheet=workbook.worksheets[0]
sheet=workbook["Sheet1"]

rows=5
cols=4

print("No of Rows: ",rows)
print("No of Columns :",cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        #sheet.cell(row=r,column=c).value="welcome"
        sheet.cell(row=r, column=c).value = "welcome"+str(random.randrange(100,400))


workbook.save(path)
