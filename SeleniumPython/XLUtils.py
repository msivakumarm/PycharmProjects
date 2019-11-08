import openpyxl


def RowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook[sheetName]
    return (sheet.max_row)

def ColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook[sheetName]
    if sheet.cell(row=rownum,column=columnno).value is not None:
        return (sheet.cell(row=rownum,column=columnno).value)
    else:
        return ""

def writeData(file,sheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet= workbook[sheetName]
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)